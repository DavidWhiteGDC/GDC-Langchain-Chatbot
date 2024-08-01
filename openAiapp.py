import streamlit as st
from openpyxl import load_workbook
from langchain_openai import AzureOpenAI
from langchain.prompts import PromptTemplate
from langchain.chains import LLMChain
import os

# Fetch environment variables from Streamlit secrets
try:
    os.environ["OPENAI_API_VERSION"] = st.secrets["OPENAI_API_VERSION"]
    os.environ["AZURE_OPENAI_ENDPOINT"] = st.secrets["AZURE_OPENAI_ENDPOINT"]
    os.environ["AZURE_OPENAI_API_KEY"] = st.secrets["AZURE_OPENAI_API_KEY"]
except KeyError as e:
    st.error(f"Missing secret: {e}")
    st.stop()

def read_excel_file(uploaded_file):
    try:
        workbook = load_workbook(uploaded_file)
        worksheet = workbook["Telemarketing Inbound"]
        rows = worksheet.iter_rows(values_only=True)
        sheet_content = []

        for row in rows:
            row_data = ", ".join([str(cell) if cell is not None else "" for cell in row])
            sheet_content.append(row_data)

        return "\n".join(sheet_content)
    except Exception as ex:
        print(f"An error occurred while reading the Excel sheet: {ex}")
        return f"An error occurred while reading the Excel sheet: {ex}"

def generate_response_from_llm(query, file_content):
    # Initialize the Azure OpenAI model
    model = AzureOpenAI(deployment_name="gpt-35-test")

    prompt_template = PromptTemplate(
        input_variables=['query', 'file_content'],
        template="The following is the content of an Excel sheet:\n{file_content}\n\nBased on this content, provide an accurate response to the following query:\n{query}-Reply only with the excel scripted response"
    )

    llm_chain = LLMChain(llm=model, prompt=prompt_template, output_key="response")

    response = llm_chain({'query': query, 'file_content': file_content})

    return response

# Streamlit UI
st.title("🤖 Excel Fetch Assistant")
st.subheader("Powered by A.I.")
st.text('Upload the excel prompt file you would like the A.I assistant to gain knowledge on!')
# File uploader to select the Excel file
uploaded_file = st.sidebar.file_uploader("Choose an Excel file", type="xlsx")

# User query input
user_query = st.text_input("Enter your query based on the Excel sheet content:")

if uploaded_file is not None:
    # Read the Excel file and store the content
    file_content = read_excel_file(uploaded_file)

    # Adding a button to trigger reading the Excel file and generating a response
    if st.button("Generate Response"):
        if user_query:
            response = generate_response_from_llm(user_query, file_content)
            st.text(response['response'])
        else:
            st.error("Please enter a query to generate a response.")
